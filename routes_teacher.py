"""
교사 기능 Blueprint - 대시보드, 학생 관리, 출석 관리, 자습시간 설정
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from models import db, User, Attendance, StudyLog, Holiday, StudyPeriodSetting, StudyRoom, StudentRoom, StudyApplication, AttendanceLog, Schedule
from constants import DEFAULT_PERIODS, WEEKDAY_CODES
from day_utils import get_day_type
import settings as app_settings
from validators import validate_password
from time_utils import validate_time_str, parse_time_str
from sqlalchemy import func
from datetime import date, datetime, timedelta
import calendar
import secrets
import random
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

teacher_bp = Blueprint('teacher', __name__)

STATUS_LABELS = {
    'present':        '출석',
    'late':           '지각',
    'absent':         '결석',
    'early_leave':    '조퇴',
    'approved_leave': '출석인정',
    'after_school':   '방과후출결인정',
}

DAY_TYPE_LABELS = {
    'mon': '월요일', 'tue': '화요일', 'wed': '수요일',
    'thu': '목요일', 'fri': '금요일',
    'weekday': '평일(공통)', 'saturday': '토요일', 'holiday': '공휴일',
}

WEEKDAY_LABELS = ['월', '화', '수', '목', '금']


def get_period_settings(day_type):
    """해당 일자 유형의 자습 시간 설정 반환.
    특정 요일 설정 → 'weekday' 공통 설정 → 기본값 순으로 폴백."""
    settings = StudyPeriodSetting.query.filter_by(
        day_type=day_type, is_active=True
    ).order_by(StudyPeriodSetting.period).all()
    if settings:
        return {s.period: (s.start_time, s.end_time) for s in settings}

    fallback = 'weekday' if day_type in WEEKDAY_CODES else day_type
    if fallback != day_type:
        settings = StudyPeriodSetting.query.filter_by(
            day_type=fallback, is_active=True
        ).order_by(StudyPeriodSetting.period).all()
        if settings:
            return {s.period: (s.start_time, s.end_time) for s in settings}

    return dict(DEFAULT_PERIODS.get(fallback, {}))


@teacher_bp.before_request
@login_required
def check_teacher():
    if current_user.role != 'teacher':
        flash('교사만 접근할 수 있습니다.', 'danger')
        return redirect(url_for('auth.login'))
    if not current_user.is_approved:
        flash('계정 승인 대기 중입니다. 관리자에게 문의하세요.', 'warning')
        return render_template('pending_approval.html')


@teacher_bp.route('/')
def dashboard():
    today = date.today()

    # 전체 학생 수
    total_students = User.query.filter_by(role='student').count()

    # 오늘 출석한 학생 수 (present/late/approved_leave 기준, 중복 제거)
    today_atts = Attendance.query.filter_by(date=today).all()
    ACTIVE_STATUSES = {'present', 'late', 'approved_leave', 'after_school'}
    attended_today = len({a.user_id for a in today_atts if a.status in ACTIVE_STATUSES})

    # 학년별 학생 수
    grade_counts = {}
    for g in range(1, 4):
        grade_counts[g] = User.query.filter_by(role='student', grade=g).count()

    # 오늘 교시별 자습 시간 설정
    day_type = get_day_type(today)
    period_times = get_period_settings(day_type)

    # 오늘 출석 현황 (실제 교시 기준)
    period_attendance = {}
    for period in (period_times.keys() if period_times else range(1, 5)):
        period_attendance[period] = sum(
            1 for a in today_atts if a.period == period and a.status in ACTIVE_STATUSES
        )

    return render_template('teacher/dashboard.html',
                           today=today,
                           total_students=total_students,
                           attended_today=attended_today,
                           grade_counts=grade_counts,
                           period_attendance=period_attendance,
                           period_times=period_times)


@teacher_bp.route('/students')
def students():
    # 담당 학년이 있으면 강제 적용 (다른 학년 열람 불가)
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)

    class_filter = request.args.get('class_num', type=int)
    room_filter = request.args.get('room', type=int)

    query = User.query.filter_by(role='student')
    if grade_filter:
        query = query.filter_by(grade=grade_filter)
    if class_filter:
        query = query.filter_by(class_num=class_filter)

    all_students = query.order_by(User.grade, User.class_num, User.name).all()

    # 학생별 배정 공간 및 좌석 번호
    student_room_map = {}
    student_seat_map = {}
    for sr in StudentRoom.query.all():
        student_room_map[sr.user_id] = sr.study_room_id
        student_seat_map[sr.user_id] = sr.seat_number

    # 공간 필터 적용
    if room_filter:
        student_list = [s for s in all_students if student_room_map.get(s.id) == room_filter]
    elif room_filter == 0:  # 미배정
        student_list = [s for s in all_students if s.id not in student_room_map]
    else:
        student_list = all_students

    # 자습 공간 목록
    study_rooms = StudyRoom.query.order_by(StudyRoom.order).all()
    room_name_map = {r.id: r.name for r in study_rooms}

    # 방과후 수업 스케줄 (user_id → sorted list of (day_of_week, period))
    _schedule_set = {}
    for sc in Schedule.query.all():
        _schedule_set.setdefault(sc.user_id, set()).add((sc.day_of_week, sc.period))
    schedule_map = {uid: sorted(v) for uid, v in _schedule_set.items()}

    # 평일 교시 목록 (스케줄 등록 그리드용)
    weekday_periods = sorted({
        s.period for s in StudyPeriodSetting.query.filter_by(is_active=True).all()
        if s.day_type in ('weekday', 'mon', 'tue', 'wed', 'thu', 'fri')
    }) or list(range(1, 5))

    return render_template('teacher/students.html',
                           students=student_list,
                           grade_filter=grade_filter,
                           class_filter=class_filter,
                           room_filter=room_filter,
                           study_rooms=study_rooms,
                           student_room_map=student_room_map,
                           student_seat_map=student_seat_map,
                           room_name_map=room_name_map,
                           assigned_grade=current_user.assigned_grade,
                           schedule_map=schedule_map,
                           weekday_periods=weekday_periods)


@teacher_bp.route('/students/assign_room', methods=['POST'])
def assign_room():
    """학생 자습 공간 배정"""
    user_id = request.form.get('user_id', type=int)
    room_id = request.form.get('room_id', type=int)

    if not user_id:
        flash('학생을 선택하세요.', 'danger')
        return redirect(url_for('teacher.students'))

    target_student = db.session.get(User, user_id)
    if not target_student or target_student.role != 'student':
        flash('유효하지 않은 학생입니다.', 'danger')
        return redirect(url_for('teacher.students'))
    if current_user.assigned_grade and target_student.grade != current_user.assigned_grade:
        flash('담당 학년의 학생만 수정할 수 있습니다.', 'danger')
        return redirect(url_for('teacher.students'))

    # 기존 배정 삭제
    StudentRoom.query.filter_by(user_id=user_id).delete()

    # 새 배정 (room_id가 있을 때만)
    if room_id:
        room = db.session.get(StudyRoom, room_id)
        if not room:
            flash('존재하지 않는 자습 공간입니다.', 'danger')
            return redirect(url_for('teacher.students'))

        # 정원 초과 검사
        if room.capacity and room.capacity > 0:
            current_count = StudentRoom.query.filter_by(study_room_id=room_id).count()
            if current_count >= room.capacity:
                flash(f'"{room.name}" 정원({room.capacity}명)이 꽉 찼습니다.', 'danger')
                return redirect(url_for('teacher.students'))

        # 남/여 정원 검사
        if target_student.gender == 'M' and room.male_capacity and room.male_capacity > 0:
            male_count = StudentRoom.query.join(User, StudentRoom.user_id == User.id)\
                .filter(StudentRoom.study_room_id == room_id, User.gender == 'M').count()
            if male_count >= room.male_capacity:
                flash(f'"{room.name}" 남학생 정원({room.male_capacity}명)이 꽉 찼습니다.', 'danger')
                return redirect(url_for('teacher.students'))
        elif target_student.gender == 'F' and room.female_capacity and room.female_capacity > 0:
            female_count = StudentRoom.query.join(User, StudentRoom.user_id == User.id)\
                .filter(StudentRoom.study_room_id == room_id, User.gender == 'F').count()
            if female_count >= room.female_capacity:
                flash(f'"{room.name}" 여학생 정원({room.female_capacity}명)이 꽉 찼습니다.', 'danger')
                return redirect(url_for('teacher.students'))

        sr = StudentRoom(user_id=user_id, study_room_id=room_id)
        db.session.add(sr)
        flash(f'자습 공간이 "{room.name}"으로 배정되었습니다.', 'success')
    else:
        flash('자습 공간 배정이 해제되었습니다.', 'info')

    db.session.commit()

    # 필터 유지
    redirect_params = {}
    if request.form.get('grade_filter'):
        redirect_params['grade'] = request.form.get('grade_filter')
    if request.form.get('class_filter'):
        redirect_params['class_num'] = request.form.get('class_filter')
    if request.form.get('room_filter'):
        redirect_params['room'] = request.form.get('room_filter')

    return redirect(url_for('teacher.students', **redirect_params))


@teacher_bp.route('/students/unassign_room', methods=['POST'])
def unassign_room():
    """학생 자습 공간 배정 해제"""
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        flash('학생을 선택하세요.', 'danger')
        return redirect(url_for('teacher.students'))

    target_student = db.session.get(User, user_id)
    if not target_student or target_student.role != 'student':
        flash('유효하지 않은 학생입니다.', 'danger')
        return redirect(url_for('teacher.students'))
    if current_user.assigned_grade and target_student.grade != current_user.assigned_grade:
        flash('담당 학년의 학생만 수정할 수 있습니다.', 'danger')
        return redirect(url_for('teacher.students'))

    deleted = StudentRoom.query.filter_by(user_id=user_id).delete()
    db.session.commit()

    if deleted:
        flash(f'{target_student.name} 학생의 자습 공간 배정이 해제되었습니다.', 'info')
    else:
        flash('배정 정보가 없습니다.', 'warning')

    redirect_params = {}
    for key in ('grade', 'class_num', 'room'):
        val = request.form.get(f'{key}_filter')
        if val:
            redirect_params[key] = val

    return redirect(url_for('teacher.students', **redirect_params))


@teacher_bp.route('/students/save_schedule', methods=['POST'])
def save_schedule():
    """학생 방과후 수업 요일·교시 저장"""
    user_id = request.form.get('user_id', type=int)
    if user_id is None:
        flash('올바른 입력값이 아닙니다.', 'danger')
        return redirect(url_for('teacher.students'))

    target = db.session.get(User, user_id)
    if not target or target.role != 'student':
        flash('유효하지 않은 학생입니다.', 'danger')
        return redirect(url_for('teacher.students'))
    if current_user.assigned_grade and target.grade != current_user.assigned_grade:
        flash('담당 학년의 학생만 수정할 수 있습니다.', 'danger')
        return redirect(url_for('teacher.students'))

    # 요일별 활성 교시 화이트리스트 구성
    # day_type 우선순위: 특정 요일(mon~fri) > weekday 공통
    DAY_TYPE_MAP = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri'}
    all_settings = StudyPeriodSetting.query.filter_by(is_active=True).all()

    def valid_periods_for_day(day_index):
        """해당 요일에 활성화된 교시 집합. 설정 없으면 None (검증 스킵)."""
        day_code = DAY_TYPE_MAP[day_index]
        specific = {s.period for s in all_settings if s.day_type == day_code}
        if specific:
            return specific
        weekday = {s.period for s in all_settings if s.day_type == 'weekday'}
        if weekday:
            return weekday
        return None  # 설정 자체가 없으면 검증 불가 → 거부

    # 유효한 항목 먼저 파싱 (삭제 전에 검증해서 조용한 데이터 소실 방지)
    new_schedules = []
    for item in request.form.getlist('sch'):
        parts = item.split('_')
        if len(parts) != 2:
            continue
        try:
            day, period = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        if not (0 <= day <= 4):
            continue
        allowed = valid_periods_for_day(day)
        if allowed is None or period not in allowed:
            continue
        new_schedules.append((day, period))

    # 기존 스케줄 삭제 후 유효한 항목만 저장
    Schedule.query.filter_by(user_id=user_id).delete()
    for day, period in new_schedules:
        db.session.add(Schedule(
            user_id=user_id,
            day_of_week=day,
            period=period,
            subject='방과후수업',
        ))
    db.session.commit()

    if new_schedules:
        flash(f'{target.name} 방과후 수업 {len(new_schedules)}건이 저장되었습니다.', 'success')
    else:
        flash(f'{target.name} 방과후 수업 일정이 초기화되었습니다. (선택 없음 또는 활성 교시 없음)', 'info')

    redirect_params = {}
    for key in ('grade', 'class_num', 'room'):
        val = request.form.get(f'{key}_filter')
        if val:
            redirect_params[key] = val
    return redirect(url_for('teacher.students', **redirect_params))


@teacher_bp.route('/attendance/after_school', methods=['POST'])
def after_school_attendance():
    """방과후 수업 참여 학생 자동 출석 처리"""
    view_date_str = request.form.get('date', '')
    try:
        view_date = date.fromisoformat(view_date_str) if view_date_str else date.today()
    except ValueError:
        view_date = date.today()

    day_of_week = view_date.weekday()  # 0=월 ... 4=금, 5=토, 6=일
    if day_of_week > 4:
        flash('토요일·일요일에는 방과후 자동출석을 처리할 수 없습니다.', 'warning')
        return redirect(url_for('teacher.attendance_view', date=view_date.isoformat()))

    schedules = Schedule.query.filter_by(day_of_week=day_of_week).all()

    # 담당 학년 필터 적용
    if current_user.assigned_grade:
        allowed_ids = {u.id for u in User.query.filter_by(
            role='student', grade=current_user.assigned_grade).all()}
        schedules = [sc for sc in schedules if sc.user_id in allowed_ids]

    if not schedules:
        flash('오늘 방과후 수업 일정이 등록된 학생이 없습니다.', 'info')
        return redirect(url_for('teacher.attendance_view', date=view_date.isoformat()))

    processed = 0
    for sc in schedules:
        # 신청 없으면 자동 생성
        if not StudyApplication.query.filter_by(
                user_id=sc.user_id, date=view_date, period=sc.period).first():
            db.session.add(StudyApplication(
                user_id=sc.user_id, date=view_date, period=sc.period))

        # 기존 출결 여부와 관계없이 after_school 처리
        existing = Attendance.query.filter_by(
                user_id=sc.user_id, date=view_date, period=sc.period).first()
        if existing:
            # 이미 after_school이면 건너뜀
            if existing.status != 'after_school':
                old_status = existing.status
                existing.status = 'after_school'
                db.session.flush()
                db.session.add(AttendanceLog(
                    attendance_id=existing.id,
                    changed_by=current_user.id,
                    old_status=old_status,
                    new_status='after_school',
                    note='방과후자동출석',
                ))
                processed += 1
        else:
            sr = StudentRoom.query.filter_by(user_id=sc.user_id).first()
            att = Attendance(
                user_id=sc.user_id, date=view_date, period=sc.period,
                status='after_school',
                study_room_id=sr.study_room_id if sr else None,
            )
            db.session.add(att)
            db.session.flush()
            db.session.add(AttendanceLog(
                attendance_id=att.id,
                changed_by=current_user.id,
                old_status=None,
                new_status='after_school',
                note='방과후자동출석',
            ))
            processed += 1

    db.session.commit()
    flash(
        f'방과후 자동출석 완료: {processed}명 출석 처리됨' if processed
        else '처리할 학생이 없습니다. (이미 처리됨 또는 일정 없음)',
        'success' if processed else 'info',
    )

    redirect_params = {'date': view_date.isoformat()}
    for key in ('grade_filter', 'room_filter', 'is_saturday', 'is_holiday'):
        val = request.form.get(key)
        if val:
            redirect_params[key.replace('_filter', '')] = val
    return redirect(url_for('teacher.attendance_view', **redirect_params))


@teacher_bp.route('/attendance')
def attendance_view():
    view_date_str = request.args.get('date', '')
    view_date = date.today()
    if view_date_str:
        try:
            view_date = date.fromisoformat(view_date_str)
        except ValueError:
            pass

    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)
    room_filter = request.args.get('room', type=int)

    # 토요일/공휴일 체크박스 (URL 파라미터로 받음)
    is_saturday = request.args.get('is_saturday') == 'on'
    is_holiday = request.args.get('is_holiday') == 'on'

    # 일자 유형 결정 (체크박스 우선, 아니면 자동 판별)
    if is_holiday:
        day_type = 'holiday'
    elif is_saturday:
        day_type = 'saturday'
    else:
        day_type = get_day_type(view_date)

    # 해당 일자 유형의 자습 시간 설정
    period_times = get_period_settings(day_type)

    # 자습 공간 목록
    study_rooms = StudyRoom.query.filter_by(is_active=True).order_by(StudyRoom.order).all()

    # 학생별 배정 공간
    student_room_map = {}
    for sr in StudentRoom.query.all():
        student_room_map[sr.user_id] = sr.study_room_id

    # 학생 목록 (공간 필터 적용)
    student_query = User.query.filter_by(role='student')
    if grade_filter:
        student_query = student_query.filter_by(grade=grade_filter)

    all_students = student_query.order_by(User.grade, User.class_num, User.name).all()

    # 공간 필터 적용
    if room_filter:
        student_list = [s for s in all_students if student_room_map.get(s.id) == room_filter]
    else:
        student_list = all_students

    # 해당 날짜 출석 데이터
    att_data = {}
    all_atts_today = Attendance.query.filter_by(date=view_date).all()
    for att in all_atts_today:
        att_data[(att.user_id, att.period)] = att

    # 해당 날짜 신청 데이터
    app_data = {}
    for app in StudyApplication.query.filter_by(date=view_date).all():
        app_data[(app.user_id, app.period)] = app

    # 해당 날짜 출결 수정 이력
    att_logs = (AttendanceLog.query
                .join(Attendance)
                .filter(Attendance.date == view_date)
                .order_by(AttendanceLog.changed_at.desc())
                .limit(100).all())

    # ── 출석 현황 요약 (필터 무관, 당일 전체 기준) ──
    _ACTIVE = {'present', 'late', 'approved_leave', 'after_school'}
    # 오늘 활성 출석이 1건 이상 있는 학생 ID
    active_user_ids = {a.user_id for a in all_atts_today if a.status in _ACTIVE}

    # 학년별 요약: {grade: {'total': N, 'present': N}}
    _all_students = User.query.filter_by(role='student').all()
    grade_summary = {}
    for s in _all_students:
        g = s.grade or 0
        if g not in grade_summary:
            grade_summary[g] = {'total': 0, 'present': 0}
        grade_summary[g]['total'] += 1
        if s.id in active_user_ids:
            grade_summary[g]['present'] += 1

    # 자습실별 요약: room_id → {'name': ..., 'present': N, 'capacity': N}
    room_summary = {r.id: {'name': r.name, 'present': 0, 'capacity': r.capacity}
                    for r in study_rooms}
    for uid in active_user_ids:
        rid = student_room_map.get(uid)
        if rid and rid in room_summary:
            room_summary[rid]['present'] += 1

    # 전체 합계
    total_students = len(_all_students)
    total_present  = len(active_user_ids)

    # 좌석 배치 현황용 데이터 (교시별 AJAX에서 사용)
    rooms_with_seats = []
    for room in study_rooms:
        assigned = (StudentRoom.query
                    .filter_by(study_room_id=room.id)
                    .join(User, StudentRoom.user_id == User.id)
                    .order_by(StudentRoom.seat_number)
                    .all())
        rooms_with_seats.append((room, assigned))

    all_active_periods = sorted({
        s.period for s in StudyPeriodSetting.query.filter_by(is_active=True).all()
    })

    # 현재 시간에 맞는 기본 교시 결정 (오늘 화면인 경우에만 적용)
    now_str = datetime.now().strftime('%H:%M')
    sorted_pts = sorted(period_times.items())
    default_period = None
    if sorted_pts:
        for p, (s, e) in sorted_pts:
            if s <= now_str <= e:
                default_period = p
                break
        if default_period is None:
            for p, (s, e) in sorted_pts:
                if now_str < s:
                    default_period = p
                    break
        if default_period is None:
            default_period = sorted_pts[-1][0]

    return render_template('teacher/attendance.html',
                           view_date=view_date,
                           students=student_list,
                           att_data=att_data,
                           app_data=app_data,
                           status_labels=STATUS_LABELS,
                           grade_filter=grade_filter,
                           room_filter=room_filter,
                           day_type=day_type,
                           day_type_label=DAY_TYPE_LABELS.get(day_type, '평일'),
                           period_times=period_times,
                           is_saturday=is_saturday,
                           is_holiday=is_holiday,
                           study_rooms=study_rooms,
                           student_room_map=student_room_map,
                           att_logs=att_logs,
                           grade_summary=grade_summary,
                           room_summary=room_summary,
                           total_students=total_students,
                           total_present=total_present,
                           rooms_with_seats=rooms_with_seats,
                           all_active_periods=all_active_periods,
                           default_period=default_period)


@teacher_bp.route('/attendance/update', methods=['POST'])
def attendance_update():
    user_id = request.form.get('user_id', type=int)
    period = request.form.get('period', type=int)
    new_status = request.form.get('status', '')
    view_date_str = request.form.get('date', '')
    early_leave_note = request.form.get('early_leave_note', '').strip()

    view_date = date.today()
    if view_date_str:
        try:
            view_date = date.fromisoformat(view_date_str)
        except ValueError:
            pass

    if user_id is None or period is None or new_status not in STATUS_LABELS:
        flash('올바른 입력값이 아닙니다.', 'danger')
        return redirect(url_for('teacher.attendance_view', date=view_date.isoformat()))

    target_student = db.session.get(User, user_id)
    if not target_student or target_student.role != 'student':
        flash('유효하지 않은 학생입니다.', 'danger')
        return redirect(url_for('teacher.attendance_view', date=view_date.isoformat()))
    if current_user.assigned_grade and target_student.grade != current_user.assigned_grade:
        flash('담당 학년의 학생만 처리할 수 있습니다.', 'danger')
        return redirect(url_for('teacher.attendance_view', date=view_date.isoformat()))

    att = Attendance.query.filter_by(
        user_id=user_id, date=view_date, period=period
    ).first()

    # 기존 상태 저장
    old_status = att.status if att else None

    if att:
        att.status = new_status
        # study_room_id는 기존 기록 유지 — 과거 자습실 이력 보존
        # 사유 메모는 항상 새 값으로 덮어씀 — 빈 값이면 None.
        # (기존 코드는 빈 사유면 옛 메모를 유지해, 상태 변경 후 무관한 메모가 남는 버그 있었음)
        if new_status in ('early_leave', 'approved_leave', 'after_school'):
            att.early_leave_note = early_leave_note or None
        else:
            att.early_leave_note = None
    else:
        # 새 출결 생성 시에만 현재 배정 공간 참조
        student_room = StudentRoom.query.filter_by(user_id=user_id).first()
        att = Attendance(
            user_id=user_id,
            date=view_date,
            period=period,
            status=new_status,
            study_room_id=student_room.study_room_id if student_room else None,
            early_leave_note=early_leave_note or None,
        )
        db.session.add(att)

    db.session.flush()  # att.id 확보

    # 출결 이력 기록
    if old_status != new_status:
        note_text = '수동처리'
        if new_status == 'approved_leave':
            note_text = f'출석인정({early_leave_note})' if early_leave_note else '출석인정'
        elif new_status == 'after_school':
            note_text = f'방과후출결인정({early_leave_note})' if early_leave_note else '방과후출결인정'
        log = AttendanceLog(
            attendance_id=att.id,
            changed_by=current_user.id,
            old_status=old_status,
            new_status=new_status,
            note=note_text,
        )
        db.session.add(log)

    db.session.commit()

    flash(f'{target_student.name} {period}교시: {STATUS_LABELS[new_status]}', 'success')

    # 필터 상태 유지
    redirect_params = {'date': view_date.isoformat()}
    if request.form.get('grade_filter'):
        redirect_params['grade'] = request.form.get('grade_filter')
    if request.form.get('room_filter'):
        redirect_params['room'] = request.form.get('room_filter')
    if request.form.get('is_saturday') == 'on':
        redirect_params['is_saturday'] = 'on'
    if request.form.get('is_holiday') == 'on':
        redirect_params['is_holiday'] = 'on'

    return redirect(url_for('teacher.attendance_view', **redirect_params))


# ========== 자습 신청 현황 ==========

@teacher_bp.route('/applications')
def applications():
    """자습 신청 현황 조회"""
    # 날짜 파라미터
    view_date_str = request.args.get('date', '')
    view_date = date.today()
    if view_date_str:
        try:
            view_date = date.fromisoformat(view_date_str)
        except ValueError:
            pass

    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)
    room_filter = request.args.get('room', type=int)

    # 일자 유형
    day_type = get_day_type(view_date)
    period_times = get_period_settings(day_type)

    # 자습 공간 목록
    study_rooms = StudyRoom.query.filter_by(is_active=True).order_by(StudyRoom.order).all()

    # 학생별 배정 공간
    student_room_map = {}
    for sr in StudentRoom.query.all():
        student_room_map[sr.user_id] = sr.study_room_id

    # 학생 목록
    student_query = User.query.filter_by(role='student')
    if grade_filter:
        student_query = student_query.filter_by(grade=grade_filter)

    all_students = student_query.order_by(User.grade, User.class_num, User.name).all()

    # 공간 필터 적용
    if room_filter:
        student_list = [s for s in all_students if student_room_map.get(s.id) == room_filter]
    else:
        student_list = all_students

    # 해당 날짜 신청 데이터
    applications = StudyApplication.query.filter_by(date=view_date).all()
    app_data = {}
    for app in applications:
        app_data[(app.user_id, app.period)] = app

    # 교시별 신청 통계 (화면 필터 기준)
    filtered_ids = {s.id for s in student_list}
    period_stats = {}
    for period in period_times.keys():
        if filtered_ids:
            count = StudyApplication.query.filter(
                StudyApplication.date == view_date,
                StudyApplication.period == period,
                StudyApplication.user_id.in_(filtered_ids)
            ).count()
        else:
            count = 0
        period_stats[period] = count

    return render_template('teacher/applications.html',
                           view_date=view_date,
                           students=student_list,
                           app_data=app_data,
                           grade_filter=grade_filter,
                           room_filter=room_filter,
                           day_type=day_type,
                           day_type_label=DAY_TYPE_LABELS.get(day_type, '평일'),
                           period_times=period_times,
                           period_stats=period_stats,
                           study_rooms=study_rooms,
                           student_room_map=student_room_map)


# ========== 월별 참여 통계 ==========

@teacher_bp.route('/statistics')
def statistics():
    """월별 참여 우수자 조회"""
    today = date.today()

    # 년/월 파라미터
    year = request.args.get('year', today.year, type=int)
    month = request.args.get('month', today.month, type=int)
    if not (2000 <= year <= 2100) or not (1 <= month <= 12):
        year, month = today.year, today.month

    # 참여율 기준 (기본값은 SystemSetting에서 조회)
    min_rate = request.args.get('min_rate', app_settings.get_int('participation_rate_default', 80), type=int)

    # 정렬 기준: 'total'(총 학습시간) 또는 'period_N'(N교시 학습시간)
    sort_by = request.args.get('sort_by', 'total')

    # 학년 필터 (담당 학년이 설정된 교사는 강제 적용)
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)

    # 해당 월의 범위
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    # 학생 목록
    student_query = User.query.filter_by(role='student')
    if grade_filter:
        student_query = student_query.filter_by(grade=grade_filter)
    students = student_query.order_by(User.grade, User.class_num, User.name).all()

    student_ids = [s.id for s in students]

    # 학생×교시별 신청/출석 수 — 벌크 쿼리
    app_by_sp = {}
    att_by_sp = {}
    min_by_sp = {}  # (user_id, period) -> 총 학습 시간(분)
    if student_ids:
        for r in db.session.query(
            StudyApplication.user_id, StudyApplication.period,
            func.count().label('cnt')
        ).filter(
            StudyApplication.user_id.in_(student_ids),
            StudyApplication.date >= first_day,
            StudyApplication.date <= last_day,
        ).group_by(StudyApplication.user_id, StudyApplication.period).all():
            app_by_sp[(r.user_id, r.period)] = r.cnt

        for r in db.session.query(
            Attendance.user_id, Attendance.period,
            func.count().label('cnt')
        ).filter(
            Attendance.user_id.in_(student_ids),
            Attendance.date >= first_day,
            Attendance.date <= last_day,
            Attendance.status.in_(['present', 'late', 'approved_leave', 'after_school']),
        ).group_by(Attendance.user_id, Attendance.period).all():
            att_by_sp[(r.user_id, r.period)] = r.cnt

        # 교시별 학습 시간 벌크 쿼리 (QR 입실·퇴실 기록이 있는 출석만)
        for r in db.session.query(
            Attendance.user_id, Attendance.period, Attendance.status,
            Attendance.study_minutes, Attendance.checked_at, Attendance.checked_out_at
        ).filter(
            Attendance.user_id.in_(student_ids),
            Attendance.date >= first_day,
            Attendance.date <= last_day,
            Attendance.checked_at.isnot(None),
            Attendance.checked_out_at.isnot(None),
        ).all():
            if r.checked_out_at > r.checked_at:
                if r.study_minutes is not None:
                    mins = r.study_minutes
                elif r.status == 'early_leave':
                    mins = 0  # 퇴실 미확인 자동 조퇴 — 학습시간 없음
                else:
                    mins = int((r.checked_out_at - r.checked_at).total_seconds() // 60)
                key = (r.user_id, r.period)
                min_by_sp[key] = min_by_sp.get(key, 0) + mins

    # 이 달 신청이 있는 교시 목록
    period_nums = sorted({p for (uid, p) in app_by_sp})

    # 학생별 통계 계산
    student_stats = []
    for student in students:
        applied_count  = sum(app_by_sp.get((student.id, p), 0) for p in period_nums)
        attended_count = sum(att_by_sp.get((student.id, p), 0) for p in period_nums)
        rate = round((attended_count / applied_count) * 100, 1) if applied_count > 0 else 0

        if rate >= min_rate:
            # 교시별 학습 시간(분)
            period_minutes = {p: min_by_sp.get((student.id, p), 0) for p in period_nums}
            total_minutes = sum(period_minutes.values())
            total_hours, rem_minutes = divmod(total_minutes, 60)

            student_stats.append({
                'student': student,
                'applied': applied_count,
                'attended': attended_count,
                'rate': rate,
                'period_minutes': period_minutes,
                'total_minutes': total_minutes,
                'total_hours': total_hours,
                'rem_minutes': rem_minutes,
            })

    # 정렬: 총 학습시간 또는 특정 교시 학습시간 기준
    if sort_by.startswith('period_') and sort_by[7:].isdigit():
        sort_period = int(sort_by[7:])
        student_stats.sort(key=lambda x: (
            -x['period_minutes'].get(sort_period, 0),
            -x['total_minutes'],
            x['student'].grade, x['student'].class_num
        ))
    else:  # 'total' (기본값)
        student_stats.sort(key=lambda x: (
            -x['total_minutes'],
            -x['attended'],
            x['student'].grade, x['student'].class_num
        ))

    # 교시별 집계 통계
    period_stats = {}
    total_applied_all = total_attended_all = total_minutes_all = 0
    for p in period_nums:
        p_applied  = sum(app_by_sp.get((uid, p), 0) for uid in student_ids)
        p_attended = sum(att_by_sp.get((uid, p), 0) for uid in student_ids)
        p_rate = round(p_attended / p_applied * 100, 1) if p_applied > 0 else 0
        p_minutes = sum(min_by_sp.get((uid, p), 0) for uid in student_ids)
        p_hours, p_rem = divmod(p_minutes, 60)
        period_stats[p] = {
            'applied': p_applied, 'attended': p_attended, 'rate': p_rate,
            'minutes': p_minutes, 'hours': p_hours, 'rem': p_rem,
        }
        total_applied_all  += p_applied
        total_attended_all += p_attended
        total_minutes_all  += p_minutes
    total_rate_all = round(total_attended_all / total_applied_all * 100, 1) if total_applied_all > 0 else 0
    total_hours_all, total_rem_all = divmod(total_minutes_all, 60)

    # 이전/다음 달
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    return render_template('teacher/statistics.html',
                           year=year,
                           month=month,
                           min_rate=min_rate,
                           sort_by=sort_by,
                           grade_filter=grade_filter,
                           student_stats=student_stats,
                           total_students=len(students),
                           period_nums=period_nums,
                           period_stats=period_stats,
                           total_applied_all=total_applied_all,
                           total_attended_all=total_attended_all,
                           total_rate_all=total_rate_all,
                           total_minutes_all=total_minutes_all,
                           total_hours_all=total_hours_all,
                           total_rem_all=total_rem_all,
                           prev_year=prev_year,
                           prev_month=prev_month,
                           next_year=next_year,
                           next_month=next_month)


# ========== 자습 시간 설정 ==========

@teacher_bp.route('/settings')
def settings():
    """자습 시간 설정 페이지"""
    all_day_types = WEEKDAY_CODES + ['weekday', 'saturday', 'holiday']
    all_settings = {}
    for dt in all_day_types:
        rows = StudyPeriodSetting.query.filter_by(day_type=dt).order_by(
            StudyPeriodSetting.period).all()
        all_settings[dt] = {s.period: s for s in rows}

    holidays = Holiday.query.order_by(Holiday.date).all()
    study_rooms = StudyRoom.query.order_by(StudyRoom.order).all()

    return render_template('teacher/settings.html',
                           all_settings=all_settings,
                           default_periods=DEFAULT_PERIODS,
                           day_type_labels=DAY_TYPE_LABELS,
                           weekday_codes=WEEKDAY_CODES,
                           weekday_labels=WEEKDAY_LABELS,
                           holidays=holidays,
                           study_rooms=study_rooms)


@teacher_bp.route('/settings/periods', methods=['POST'])
def settings_periods():
    """자습 시간 설정 저장 (특정 요일 또는 공통 평일/토/공휴일)"""
    valid_types = WEEKDAY_CODES + ['weekday', 'saturday', 'holiday']
    day_type = request.form.get('day_type', '')
    if day_type not in valid_types:
        flash('올바르지 않은 일자 유형입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # 동일 설정을 함께 저장할 대상 요일 목록 (폼에서 체크박스로 선택)
    copy_to = request.form.getlist('copy_to')
    # 저장 대상: 기본 day_type + 선택한 추가 요일
    targets = [day_type] + [d for d in copy_to if d in valid_types and d != day_type]

    # 아침 자습(0교시) 포함 여부: 평일 계열만
    weekday_family = WEEKDAY_CODES + ['weekday']
    period_range = range(0, 5) if day_type in weekday_family else range(1, 5)

    rows = []
    for period in period_range:
        is_active = request.form.get(f'active_{period}') == 'on'
        start_time = request.form.get(f'start_{period}', '')
        end_time = request.form.get(f'end_{period}', '')
        if start_time and end_time:
            rows.append((period, start_time, end_time, is_active))

    for period, start_time, end_time, is_active in rows:
        ok_s, err_s = validate_time_str(start_time, f'{period}교시 시작 시각')
        if not ok_s:
            flash(err_s, 'danger')
            return redirect(url_for('teacher.settings'))
        ok_e, err_e = validate_time_str(end_time, f'{period}교시 종료 시각')
        if not ok_e:
            flash(err_e, 'danger')
            return redirect(url_for('teacher.settings'))
        if parse_time_str(start_time) >= parse_time_str(end_time):
            flash(f'{period}교시 시작 시각이 종료 시각보다 늦거나 같습니다.', 'danger')
            return redirect(url_for('teacher.settings'))

    # 교시 간 시간 겹침 검사 (활성 교시끼리만)
    parsed_rows = []
    for period, start_time, end_time, is_active in rows:
        if not is_active:
            continue
        parsed_rows.append((period, parse_time_str(start_time), parse_time_str(end_time)))
    for i in range(len(parsed_rows)):
        for j in range(i + 1, len(parsed_rows)):
            p1, s1, e1 = parsed_rows[i]
            p2, s2, e2 = parsed_rows[j]
            if s1 < e2 and s2 < e1:
                flash(f'{p1}교시와 {p2}교시 시간이 겹칩니다.', 'danger')
                return redirect(url_for('teacher.settings'))

    for target in targets:
        StudyPeriodSetting.query.filter_by(day_type=target).delete()
        for period, start_time, end_time, is_active in rows:
            db.session.add(StudyPeriodSetting(
                day_type=target,
                period=period,
                start_time=start_time,
                end_time=end_time,
                is_active=is_active,
            ))

    db.session.commit()
    saved_labels = [DAY_TYPE_LABELS.get(t, t) for t in targets]
    flash(f'{", ".join(saved_labels)} 자습 시간이 저장되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/periods/delete/<day_type>', methods=['POST'])
def settings_periods_delete(day_type):
    """특정 요일 개별 설정 삭제 (공통 설정으로 되돌리기)"""
    if day_type not in WEEKDAY_CODES:
        flash('개별 삭제는 월~금 요일만 가능합니다.', 'danger')
        return redirect(url_for('teacher.settings'))
    StudyPeriodSetting.query.filter_by(day_type=day_type).delete()
    db.session.commit()
    flash(f'{DAY_TYPE_LABELS[day_type]} 개별 설정이 삭제되었습니다. 평일(공통) 설정이 적용됩니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/holiday', methods=['POST'])
def settings_holiday():
    """공휴일 추가"""
    date_str = request.form.get('date', '')
    name = request.form.get('name', '').strip()

    if not date_str or not name:
        flash('날짜와 공휴일 이름을 입력하세요.', 'danger')
        return redirect(url_for('teacher.settings'))

    try:
        holiday_date = date.fromisoformat(date_str)
    except ValueError:
        flash('올바른 날짜 형식이 아닙니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # 중복 체크
    if Holiday.query.filter_by(date=holiday_date).first():
        flash('이미 등록된 날짜입니다.', 'warning')
        return redirect(url_for('teacher.settings'))

    holiday = Holiday(date=holiday_date, name=name)
    db.session.add(holiday)
    db.session.commit()

    flash(f'{holiday_date.strftime("%Y-%m-%d")} {name} 공휴일이 등록되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/holiday/delete/<int:holiday_id>', methods=['POST'])
def settings_holiday_delete(holiday_id):
    """공휴일 삭제"""
    holiday = db.session.get(Holiday, holiday_id)
    if holiday:
        db.session.delete(holiday)
        db.session.commit()
        flash('공휴일이 삭제되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


# ========== 자습 공간 관리 ==========

@teacher_bp.route('/settings/room', methods=['POST'])
def settings_room():
    """자습 공간 추가"""
    name = request.form.get('name', '').strip()
    capacity = request.form.get('capacity', type=int) or 0
    male_capacity = request.form.get('male_capacity', type=int) or 0
    female_capacity = request.form.get('female_capacity', type=int) or 0

    if not name:
        flash('공간 이름을 입력하세요.', 'danger')
        return redirect(url_for('teacher.settings'))

    if StudyRoom.query.filter_by(name=name).first():
        flash(f'"{name}" 이름의 자습 공간이 이미 존재합니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # 정원 유효성 검사
    if capacity < 0 or male_capacity < 0 or female_capacity < 0:
        flash('정원은 0 이상이어야 합니다.', 'danger')
        return redirect(url_for('teacher.settings'))
    if capacity > 0 and (male_capacity + female_capacity) > capacity:
        flash('남학생+여학생 정원 합계가 전체 정원을 초과할 수 없습니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # 순서 자동 설정
    max_order = db.session.query(db.func.max(StudyRoom.order)).scalar() or 0

    # QR 토큰 자동 생성
    qr_token = secrets.token_hex(16)

    room = StudyRoom(
        name=name,
        capacity=capacity,
        male_capacity=male_capacity,
        female_capacity=female_capacity,
        order=max_order + 1,
        qr_token=qr_token
    )
    db.session.add(room)
    db.session.commit()

    flash(f'자습 공간 "{name}"이(가) 추가되었습니다. QR코드가 자동 생성되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/room/toggle/<int:room_id>', methods=['POST'])
def settings_room_toggle(room_id):
    """자습 공간 활성화/비활성화"""
    room = db.session.get(StudyRoom, room_id)
    if room:
        room.is_active = not room.is_active
        db.session.commit()
        status = '활성화' if room.is_active else '비활성화'
        flash(f'자습 공간 "{room.name}"이(가) {status}되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/room/edit/<int:room_id>', methods=['POST'])
def settings_room_edit(room_id):
    """자습 공간 이름·수용인원 수정"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    name = request.form.get('name', '').strip()
    if not name:
        flash('공간 이름을 입력하세요.', 'danger')
        return redirect(url_for('teacher.settings'))

    dup = StudyRoom.query.filter_by(name=name).first()
    if dup and dup.id != room_id:
        flash(f'"{name}" 이름의 자습 공간이 이미 존재합니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    new_capacity        = request.form.get('capacity',        type=int) or 0
    new_male_capacity   = request.form.get('male_capacity',   type=int) or 0
    new_female_capacity = request.form.get('female_capacity', type=int) or 0

    # 정원 유효성 검사
    if new_capacity < 0 or new_male_capacity < 0 or new_female_capacity < 0:
        flash('정원은 0 이상이어야 합니다.', 'danger')
        return redirect(url_for('teacher.settings'))
    if new_capacity > 0 and (new_male_capacity + new_female_capacity) > new_capacity:
        flash('남학생+여학생 정원 합계가 전체 정원을 초과할 수 없습니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    room.name            = name
    room.capacity        = new_capacity
    room.male_capacity   = new_male_capacity
    room.female_capacity = new_female_capacity
    db.session.commit()
    flash(f'자습 공간 "{name}"이(가) 수정되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/room/delete/<int:room_id>', methods=['POST'])
def settings_room_delete(room_id):
    """자습 공간 삭제 (연관 출결 기록은 room 참조를 NULL로, 배정 기록은 삭제)"""
    room = db.session.get(StudyRoom, room_id)
    if room:
        name = room.name
        # 출결 기록의 study_room_id 참조를 NULL로 초기화
        Attendance.query.filter_by(study_room_id=room_id).update(
            {'study_room_id': None}, synchronize_session=False
        )
        # 학생 배정 기록 삭제
        StudentRoom.query.filter_by(study_room_id=room_id).delete(
            synchronize_session=False
        )
        db.session.delete(room)
        db.session.commit()
        flash(f'자습 공간 "{name}"이(가) 삭제되었습니다.', 'success')
    return redirect(url_for('teacher.settings'))


@teacher_bp.route('/settings/room/<int:room_id>/assign_seats', methods=['POST'])
def assign_seats(room_id):
    """자습실 학생 자리 랜덤 배치 (성별 구분)"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # 해당 공간에 배정된 학생 조회
    assigned = StudentRoom.query.filter_by(study_room_id=room_id).all()
    males   = [sr for sr in assigned if sr.user and sr.user.gender == 'M']
    females = [sr for sr in assigned if sr.user and sr.user.gender == 'F']

    warnings = []

    # 남학생 자리 배치
    if males:
        n = room.male_capacity if room.male_capacity > 0 else len(males)
        if len(males) > n:
            warnings.append(f'남학생 수({len(males)})가 남학생 정원({n})을 초과합니다.')
            n = len(males)
        seats = random.sample(range(1, n + 1), len(males))
        for sr, seat in zip(males, seats):
            sr.seat_number = seat

    # 여학생 자리 배치
    if females:
        n = room.female_capacity if room.female_capacity > 0 else len(females)
        if len(females) > n:
            warnings.append(f'여학생 수({len(females)})가 여학생 정원({n})을 초과합니다.')
            n = len(females)
        seats = random.sample(range(1, n + 1), len(females))
        for sr, seat in zip(females, seats):
            sr.seat_number = seat

    db.session.commit()

    for w in warnings:
        flash(w, 'warning')

    flash(
        f'"{room.name}" 임의 배정 완료 — '
        f'남학생 {len(males)}명 / 여학생 {len(females)}명. '
        f'배치도에서 좌석 위치를 조정하세요.',
        'success'
    )
    return redirect(url_for('teacher.seat_layout', room_id=room_id))


@teacher_bp.route('/settings/room/<int:room_id>/manual_seats', methods=['GET', 'POST'])
def manual_seats(room_id):
    """자습실 학생 수동 자리 배정"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    assigned = StudentRoom.query.filter_by(study_room_id=room_id).join(User).all()
    males   = [sr for sr in assigned if sr.user and sr.user.gender == 'M']
    females = [sr for sr in assigned if sr.user and sr.user.gender == 'F']
    male_cap   = room.male_capacity   if room.male_capacity   > 0 else max(len(males),   1)
    female_cap = room.female_capacity if room.female_capacity > 0 else max(len(females), 1)

    if request.method == 'POST':
        # 기존 seat_number 초기화
        for sr in assigned:
            sr.seat_number = None
        db.session.flush()

        # 남학생 배정 처리
        used_m = set()
        for i in range(1, male_cap + 1):
            uid = request.form.get(f'seat_m_{i}', type=int)
            if uid and uid not in used_m:
                sr = next((s for s in males if s.user_id == uid), None)
                if sr:
                    sr.seat_number = i
                    used_m.add(uid)

        # 여학생 배정 처리
        used_f = set()
        for i in range(1, female_cap + 1):
            uid = request.form.get(f'seat_f_{i}', type=int)
            if uid and uid not in used_f:
                sr = next((s for s in females if s.user_id == uid), None)
                if sr:
                    sr.seat_number = i
                    used_f.add(uid)

        db.session.commit()

        assigned_count = len(used_m) + len(used_f)
        total_count    = len(males) + len(females)
        flash(
            f'"{room.name}" 수동 배정 완료 — '
            f'{assigned_count}/{total_count}명 배정됨. '
            f'배치도에서 좌석 위치를 조정하세요.',
            'success'
        )
        return redirect(url_for('teacher.seat_layout', room_id=room_id))

    # 현재 배정 현황 (seat_number → user)
    seat_map_m = {sr.seat_number: sr.user for sr in males   if sr.seat_number}
    seat_map_f = {sr.seat_number: sr.user for sr in females if sr.seat_number}

    return render_template('teacher/manual_seats.html',
                           room=room,
                           males=males,
                           females=females,
                           seat_map_m=seat_map_m,
                           seat_map_f=seat_map_f,
                           male_cap=male_cap,
                           female_cap=female_cap)


@teacher_bp.route('/attendance/auto_process', methods=['POST'])
def attendance_auto_process():
    """신청자 중 미처리 학생 자동 지각/결석 처리"""
    view_date_str = request.form.get('date', '')
    try:
        view_date = date.fromisoformat(view_date_str) if view_date_str else date.today()
    except ValueError:
        view_date = date.today()

    # 담당 학년 교사는 자기 학년만 처리 가능 - form의 grade를 신뢰하지 않는다
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.form.get('grade', type=int)
    room_filter  = request.form.get('room',  type=int)
    # attendance_view의 토요일/공휴일 override를 자동처리에도 동일하게 반영해야
    # 화면-서버 기준 불일치를 막는다. (form의 day_type 자체는 신뢰하지 않고 서버에서 재계산)
    is_saturday = request.form.get('is_saturday') == 'on'
    is_holiday  = request.form.get('is_holiday')  == 'on'
    if is_holiday:
        day_type = 'holiday'
    elif is_saturday:
        day_type = 'saturday'
    else:
        day_type = get_day_type(view_date)
    period_times = get_period_settings(day_type)
    now_dt   = datetime.now()
    is_today = (view_date == date.today())

    processed = 0
    for period, (start_str, end_str) in period_times.items():
        if is_today:
            start_dt = datetime.combine(view_date, datetime.strptime(start_str, '%H:%M').time())
            end_dt   = datetime.combine(view_date, datetime.strptime(end_str,   '%H:%M').time())
            late_threshold = start_dt + timedelta(minutes=app_settings.get_int('late_threshold_minutes', 10))
            if now_dt < late_threshold:
                continue
            status = 'absent' if now_dt >= end_dt else 'late'
        else:
            status = 'absent'

        apps = StudyApplication.query.filter_by(date=view_date, period=period).all()
        if grade_filter or room_filter:
            grade_ids = {u.id for u in User.query.filter_by(role='student').all()
                        if not grade_filter or u.grade == grade_filter}
            if room_filter:
                room_ids = {sr.user_id for sr in StudentRoom.query.filter_by(study_room_id=room_filter).all()}
                grade_ids &= room_ids
            apps = [a for a in apps if a.user_id in grade_ids]
        for app in apps:
            existing = Attendance.query.filter_by(
                user_id=app.user_id, date=view_date, period=period).first()
            if existing:
                # 승격 케이스: 시작 직후 자동처리로 late가 박혔는데 끝까지 미입실이면 absent로 승격
                # (오늘 + 종료 시각 지남 + 아직 입실 미확인 + 현재 late 상태)
                if (is_today
                        and now_dt >= end_dt
                        and existing.status == 'late'
                        and existing.checked_at is None):
                    db.session.add(AttendanceLog(
                        attendance_id=existing.id, changed_by=current_user.id,
                        old_status='late', new_status='absent', note='자동승격(미입실)',
                    ))
                    existing.status = 'absent'
                    processed += 1
                continue
            sr = StudentRoom.query.filter_by(user_id=app.user_id).first()
            att = Attendance(
                user_id=app.user_id, date=view_date, period=period,
                status=status,
                study_room_id=sr.study_room_id if sr else None,
                checked_at=None  # QR 미스캔이므로 checked_at 없음
            )
            db.session.add(att)
            db.session.flush()
            db.session.add(AttendanceLog(
                attendance_id=att.id, changed_by=current_user.id,
                old_status=None, new_status=status, note='자동처리'
            ))
            processed += 1

    db.session.commit()

    label = '결석' if not is_today else '지각/결석'
    flash(f'자동 처리 완료: {processed}명 {label} 처리됨' if processed
          else '처리할 학생이 없습니다. (이미 처리됨 또는 신청자 없음)', 'success' if processed else 'info')

    redirect_params = {'date': view_date.isoformat()}
    for k in ('grade', 'room', 'is_saturday', 'is_holiday'):
        v = request.form.get(k)
        if v:
            redirect_params[k] = v
    return redirect(url_for('teacher.attendance_view', **redirect_params))


@teacher_bp.route('/students/<int:user_id>/report')
def student_report(user_id):
    """학생 개인 리포트 (월별 출석/신청/학습 요약)"""
    student = db.session.get(User, user_id)
    if not student or student.role != 'student':
        flash('학생을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('teacher.students'))

    if current_user.assigned_grade and student.grade != current_user.assigned_grade:
        flash('담당 학년의 학생만 조회할 수 있습니다.', 'danger')
        return redirect(url_for('teacher.students'))

    today = date.today()
    year  = request.args.get('year',  today.year,  type=int)
    month = request.args.get('month', today.month, type=int)
    if not (1 <= month <= 12) or not (2000 <= year <= 2100):
        year, month = today.year, today.month

    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])

    attendances  = Attendance.query.filter(
        Attendance.user_id == user_id,
        Attendance.date >= first_day, Attendance.date <= last_day
    ).order_by(Attendance.date, Attendance.period).all()

    applications = StudyApplication.query.filter(
        StudyApplication.user_id == user_id,
        StudyApplication.date >= first_day, StudyApplication.date <= last_day
    ).order_by(StudyApplication.date, StudyApplication.period).all()

    study_logs = StudyLog.query.filter(
        StudyLog.user_id == user_id,
        StudyLog.date >= first_day, StudyLog.date <= last_day
    ).order_by(StudyLog.date.desc()).all()

    # 각 카운트는 상호 배타적이어야 한다 (rate에서 이중 계산 방지).
    # 템플릿이 present_count를 "출석" 라벨로 표시하므로 present_count는 좁게 유지.
    # approved_count는 출석인정 + 방과후출결인정을 합산 (둘 다 사실상 출석 처리).
    present_count  = sum(1 for a in attendances if a.status == 'present')
    late_count     = sum(1 for a in attendances if a.status == 'late')
    absent_count   = sum(1 for a in attendances if a.status == 'absent')
    approved_count = sum(1 for a in attendances if a.status in ('approved_leave', 'after_school'))
    applied_count  = len(applications)
    rate = round((present_count + late_count + approved_count) / applied_count * 100, 1) if applied_count > 0 else 0
    total_study_min = sum(l.duration for l in study_logs)

    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year  = year if month < 12 else year + 1

    return render_template('teacher/student_report.html',
                           student=student,
                           year=year, month=month,
                           attendances=attendances,
                           applications=applications,
                           study_logs=study_logs,
                           present_count=present_count,
                           late_count=late_count,
                           absent_count=absent_count,
                           approved_count=approved_count,
                           applied_count=applied_count,
                           rate=rate,
                           total_study_min=total_study_min,
                           prev_year=prev_year, prev_month=prev_month,
                           next_year=next_year, next_month=next_month,
                           status_labels=STATUS_LABELS)


@teacher_bp.route('/export/attendance')
def export_attendance():
    """출석부 Excel 다운로드"""
    view_date_str = request.args.get('date', date.today().isoformat())
    try:
        view_date = date.fromisoformat(view_date_str)
    except ValueError:
        view_date = date.today()

    is_saturday = request.args.get('is_saturday') == 'on'
    is_holiday  = request.args.get('is_holiday')  == 'on'
    if is_holiday:
        day_type = 'holiday'
    elif is_saturday:
        day_type = 'saturday'
    else:
        day_type = get_day_type(view_date)

    period_times = get_period_settings(day_type)
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)
    room_filter  = request.args.get('room', type=int)

    q = User.query.filter_by(role='student')
    if grade_filter:
        q = q.filter_by(grade=grade_filter)
    all_students = q.order_by(User.grade, User.class_num, User.name).all()

    sr_map = {sr.user_id: sr.study_room_id for sr in StudentRoom.query.all()}
    if room_filter:
        students = [s for s in all_students if sr_map.get(s.id) == room_filter]
    else:
        students = all_students

    att_data = {(a.user_id, a.period): a
                for a in Attendance.query.filter_by(date=view_date).all()}
    room_map = {r.id: r.name for r in StudyRoom.query.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = view_date.strftime('%Y-%m-%d')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hfill  = PatternFill(fill_type='solid', fgColor='4472C4')

    periods_sorted = sorted(period_times.keys())
    total_cols = 5 + len(periods_sorted) * 3

    # 1행: 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
                value=f'{view_date.strftime("%Y년 %m월 %d일")} 출석부')
    t.font = Font(bold=True, size=14)
    t.alignment = center

    # 2행: 단일 헤더 행 (병합 없음)
    headers = ['이름', '학번', '학년', '반', '자습공간']
    for p in periods_sorted:
        st_time, et_time = period_times[p]
        headers.append(f'{p}교시\n상태\n({st_time}~{et_time})')
        headers.append(f'{p}교시\n입실')
        headers.append(f'{p}교시\n퇴실')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = center
        c.fill = hfill
    ws.row_dimensions[2].height = 42

    status_text  = {'present': '출석', 'late': '지각', 'absent': '결석',
                    'early_leave': '조퇴', 'approved_leave': '출석인정',
                    'after_school': '방과후출결인정'}
    status_color = {'present': 'C6EFCE', 'late': 'FFEB9C', 'absent': 'FFC7CE',
                    'early_leave': 'F4CCCC', 'approved_leave': 'CFE2F3',
                    'after_school': 'D9C8F5'}

    for ri, s in enumerate(students, 3):
        ws.cell(ri, 1, s.name)
        ws.cell(ri, 2, s.student_id or '')
        ws.cell(ri, 3, f'{s.grade}학년')
        ws.cell(ri, 4, f'{s.class_num}반')
        att_room_id = next(
            (att_data[(s.id, p)].study_room_id
             for p in periods_sorted
             if (s.id, p) in att_data and att_data[(s.id, p)].study_room_id),
            sr_map.get(s.id)
        )
        ws.cell(ri, 5, room_map.get(att_room_id, '미배정') if att_room_id else '미배정')

        for i, p in enumerate(periods_sorted):
            col_start = 6 + i * 3
            att = att_data.get((s.id, p))
            st  = att.status if att else None

            status_cell = ws.cell(ri, col_start, status_text.get(st, '-'))
            status_cell.alignment = center
            if st in status_color:
                fill = PatternFill(fill_type='solid', fgColor=status_color[st])
                status_cell.fill = fill

            ws.cell(ri, col_start + 1,
                    att.checked_at.strftime('%H:%M') if att and att.checked_at else '')
            ws.cell(ri, col_start + 2,
                    att.checked_out_at.strftime('%H:%M') if att and att.checked_out_at else '')

    # 열 너비
    for col, w in zip('ABCDE', [12, 10, 8, 6, 15]):
        ws.column_dimensions[col].width = w
    for i in range(len(periods_sorted)):
        col_start = 6 + i * 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_start)].width = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_start + 1)].width = 8
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_start + 2)].width = 8

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'출석부_{view_date.strftime("%Y%m%d")}.xlsx')


@teacher_bp.route('/export/attendance_range')
def export_attendance_range():
    """기간별 전체 출결 현황 Excel 다운로드 (자습실 구분 없음)"""
    try:
        date_from = date.fromisoformat(request.args.get('date_from', ''))
        date_to   = date.fromisoformat(request.args.get('date_to', ''))
    except ValueError:
        flash('날짜 형식이 올바르지 않습니다.', 'danger')
        return redirect(url_for('teacher.attendance_view'))

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    # 선택된 교시 파싱
    selected_periods = sorted(
        int(p) for p in request.args.getlist('periods') if p.isdigit()
    )
    if not selected_periods:
        flash('교시를 하나 이상 선택하세요.', 'warning')
        return redirect(url_for('teacher.attendance_view'))

    # 학년 필터 (담당 학년 교사는 자기 학년 강제, 그 외 0/미지정은 전체)
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)

    # 날짜 목록 (토·일 제외 옵션 없이 전부 포함)
    delta = (date_to - date_from).days + 1
    date_list = [date_from + timedelta(days=i) for i in range(delta)]

    # 학생 목록 (학년·반·이름 순)
    sq = User.query.filter_by(role='student')
    if grade_filter:
        sq = sq.filter_by(grade=grade_filter)
    students = sq.order_by(User.grade, User.class_num, User.name).all()

    # 자습실 맵
    room_map = {r.id: r.name for r in StudyRoom.query.all()}
    sr_map   = {sr.user_id: sr.study_room_id for sr in StudentRoom.query.all()}

    # 출결 데이터 한 번에 조회
    att_records = (Attendance.query
                   .filter(Attendance.date >= date_from,
                           Attendance.date <= date_to,
                           Attendance.period.in_(selected_periods))
                   .all())
    att_map = {(a.user_id, a.date, a.period): a for a in att_records}

    status_text  = {'present': '출석', 'late': '지각', 'absent': '결석',
                    'early_leave': '조퇴', 'approved_leave': '출석인정',
                    'after_school': '방과후출결인정'}
    status_color = {'present': 'C6EFCE', 'late': 'FFEB9C', 'absent': 'FFC7CE',
                    'early_leave': 'F4CCCC', 'approved_leave': 'CFE2F3',
                    'after_school': 'D9C8F5'}

    from openpyxl.styles import Border, Side
    thick  = Side(style='medium', color='4472C4')
    thin   = Side(style='thin',   color='BFBFBF')
    medium = Side(style='medium', color='000000')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '기간별출결'

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hfill  = PatternFill(fill_type='solid', fgColor='4472C4')
    sfill  = PatternFill(fill_type='solid', fgColor='FFF2CC')
    wfill  = PatternFill(fill_type='solid', fgColor='1F3864')

    fixed_headers = ['이름', '학번', '학년', '반', '자습공간']
    FIXED = len(fixed_headers)

    WEEKDAY_KO  = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    grade_label = f'{grade_filter}학년 ' if grade_filter else '전체 학년 '

    # date_list를 ISO 캘린더 주 단위로 묶기
    weeks: list = []
    cur_week_dates: list = []
    for d in date_list:
        if cur_week_dates and d.isocalendar()[:2] != cur_week_dates[0].isocalendar()[:2]:
            weeks.append(cur_week_dates)
            cur_week_dates = []
        cur_week_dates.append(d)
    if cur_week_dates:
        weeks.append(cur_week_dates)

    # 고정 열 너비
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 7
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 12

    cur_row  = 1
    week_num = 0

    for week_dates in weeks:
        week_num += 1
        N       = len(selected_periods)
        week_dp = [(d, p) for d in week_dates for p in selected_periods]
        total_cols = FIXED + len(week_dp)

        # ── 주차 타이틀 행 ──
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=total_cols)
        w_start = f'{week_dates[0].strftime("%m/%d")}({WEEKDAY_KO[week_dates[0].weekday()]})'
        w_end   = f'{week_dates[-1].strftime("%m/%d")}({WEEKDAY_KO[week_dates[-1].weekday()]})'
        t = ws.cell(row=cur_row, column=1,
                    value=f'{grade_label}{week_num}주차  {w_start} ~ {w_end}  출결현황'
                          f'  ({", ".join(str(p)+"교시" for p in selected_periods)})')
        t.font      = Font(bold=True, size=12, color='FFFFFF')
        t.alignment = center
        t.fill      = wfill
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # ── 날짜 병합 헤더 행 ──
        ws.row_dimensions[cur_row].height = 18
        for col, h in enumerate(fixed_headers, 1):
            ws.merge_cells(start_row=cur_row, start_column=col,
                           end_row=cur_row + 1, end_column=col)
            c = ws.cell(row=cur_row, column=col, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.alignment = center
            c.fill      = hfill
        col_idx = FIXED + 1
        for d in week_dates:
            ws.merge_cells(start_row=cur_row, start_column=col_idx,
                           end_row=cur_row, end_column=col_idx + N - 1)
            dc = ws.cell(row=cur_row, column=col_idx,
                         value=f'{d.strftime("%m/%d")}({WEEKDAY_KO[d.weekday()]})')
            dc.font      = Font(bold=True, color='FFFFFF')
            dc.alignment = center
            dc.fill      = hfill
            col_idx += N
        cur_row += 1

        # ── 교시 서브헤더 행 ──
        ws.row_dimensions[cur_row].height = 16
        for col in range(1, FIXED + 1):
            c = ws.cell(row=cur_row, column=col)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = hfill
            c.alignment = center
        for i, (_, p) in enumerate(week_dp):
            c = ws.cell(row=cur_row, column=FIXED + 1 + i, value=f'{p}교시')
            c.font      = Font(bold=True, color='FFFFFF')
            c.alignment = center
            c.fill      = hfill
        cur_row += 1

        # ── 감독교사 서명 행 ──
        ws.row_dimensions[cur_row].height = 45
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=FIXED)
        lc = ws.cell(row=cur_row, column=1, value='감독교사 서명')
        lc.font      = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal='center', vertical='center')
        lc.fill      = sfill
        lc.border    = Border(top=medium, bottom=medium, left=medium, right=thin)
        for i, (_, p) in enumerate(week_dp):
            col      = FIXED + 1 + i
            is_first = (p == selected_periods[0])
            is_last  = (p == selected_periods[-1])
            label    = f'{p}교시\n(서명)' if N > 1 else '(서명)'
            sc = ws.cell(row=cur_row, column=col, value=label)
            sc.font      = Font(size=8, color='7F7F7F')
            sc.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            sc.fill      = sfill
            sc.border    = Border(
                top=medium, bottom=medium,
                left=medium if is_first else thin,
                right=medium if is_last else thin,
            )
        cur_row += 1

        # ── 학생 데이터 행 ──
        data_start = cur_row
        for ri_off, s in enumerate(students):
            ri = cur_row + ri_off
            room_id = sr_map.get(s.id)
            ws.cell(ri, 1, s.name)
            ws.cell(ri, 2, s.student_id or '')
            ws.cell(ri, 3, f'{s.grade}학년')
            ws.cell(ri, 4, f'{s.class_num}반')
            ws.cell(ri, 5, room_map.get(room_id, '미배정') if room_id else '미배정')
            for col in range(1, FIXED + 1):
                ws.cell(ri, col).alignment = Alignment(horizontal='center', vertical='center')
            for i, (d, p) in enumerate(week_dp):
                att  = att_map.get((s.id, d, p))
                st   = att.status if att else None
                cell = ws.cell(ri, FIXED + 1 + i, status_text.get(st, '-'))
                cell.alignment = center
                if st in status_color:
                    cell.fill = PatternFill(fill_type='solid', fgColor=status_color[st])
        cur_row += len(students)

        # ── 날짜 경계 세로선 (날짜헤더·교시헤더·데이터 행, 서명 행 제외) ──
        sign_row = data_start - 1
        for ri in range(data_start - 3, cur_row):
            if ri == sign_row:
                continue
            for i, (_, p) in enumerate(week_dp):
                cell = ws.cell(ri, FIXED + 1 + i)
                is_last_p  = (p == selected_periods[-1])
                not_last_c = (FIXED + 1 + i < total_cols)
                right_side = thick if (is_last_p and not_last_c) else thin
                cell.border = Border(right=right_side, top=thin, bottom=thin, left=thin)

        # 교시 열 너비
        for i in range(len(week_dp)):
            col_letter = openpyxl.utils.get_column_letter(FIXED + 1 + i)
            ws.column_dimensions[col_letter].width = 8

        # 주 사이 구분 빈 행
        cur_row += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    grade_suffix = f'_{grade_filter}학년' if grade_filter else ''
    fname = f'출결현황{grade_suffix}_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx'
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname)


@teacher_bp.route('/export/statistics')
def export_statistics():
    """참여 통계 Excel 다운로드"""
    today = date.today()
    year      = request.args.get('year',     today.year,  type=int)
    month     = request.args.get('month',    today.month, type=int)
    min_rate  = request.args.get('min_rate', 0,           type=int)
    if current_user.assigned_grade:
        grade_filter = current_user.assigned_grade
    else:
        grade_filter = request.args.get('grade', type=int)

    if not (1 <= month <= 12) or not (2000 <= year <= 2100):
        year, month = today.year, today.month
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])

    q = User.query.filter_by(role='student')
    if grade_filter:
        q = q.filter_by(grade=grade_filter)
    students = q.order_by(User.grade, User.class_num, User.name).all()

    rows = []
    for s in students:
        applied  = StudyApplication.query.filter(
            StudyApplication.user_id == s.id,
            StudyApplication.date.between(first_day, last_day)
        ).count()
        attended = Attendance.query.filter(
            Attendance.user_id == s.id,
            Attendance.date.between(first_day, last_day),
            Attendance.status.in_(['present', 'late', 'approved_leave', 'after_school'])
        ).count()
        att_records = Attendance.query.filter(
            Attendance.user_id == s.id,
            Attendance.date.between(first_day, last_day),
            Attendance.checked_out_at.isnot(None),
            Attendance.checked_at.isnot(None)
        ).all()
        total_minutes = sum(
            a.study_minutes if a.study_minutes is not None
            else (0 if a.status == 'early_leave'
                  else int((a.checked_out_at - a.checked_at).total_seconds() // 60))
            for a in att_records
            if a.checked_out_at > a.checked_at
        )
        rate = round(attended / applied * 100, 1) if applied > 0 else 0
        if rate >= min_rate:
            rows.append((s, applied, attended, rate, total_minutes))
    rows.sort(key=lambda x: (-x[3], -x[2]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{year}년{month}월'

    center = Alignment(horizontal='center', vertical='center')
    hfill  = PatternFill(fill_type='solid', fgColor='70AD47')

    ws.merge_cells('A1:I1')
    t = ws['A1']
    t.value     = f'{year}년 {month}월 참여 통계 (기준: {min_rate}% 이상)'
    t.font      = Font(bold=True, size=14)
    t.alignment = center

    for col, h in enumerate(['순위','이름','학번','학년','반','신청','출석','참여율','총 자습 시간'], 1):
        c = ws.cell(2, col, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = center
        c.fill = hfill

    for i, (s, applied, attended, rate, total_minutes) in enumerate(rows, 1):
        color = 'C6EFCE' if rate >= 90 else ('DDEBF7' if rate >= 80 else 'FFEB9C')
        h, m = divmod(total_minutes, 60)
        time_str = f'{h}시간 {m}분' if total_minutes > 0 else '-'
        for col, val in enumerate(
            [i, s.name, s.student_id or '', f'{s.grade}학년', f'{s.class_num}반',
             applied, attended, f'{rate}%', time_str], 1):
            c = ws.cell(i + 2, col, val)
            c.alignment = center
            if col == 8:
                c.fill = PatternFill(fill_type='solid', fgColor=color)

    for col, w in zip('ABCDEFGHI', [8, 12, 10, 8, 6, 8, 8, 10, 14]):
        ws.column_dimensions[col].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'참여통계_{year}년{month}월.xlsx')


# ========== QR코드 출석 체크 ==========

@teacher_bp.route('/qr/<int:room_id>')
def qr_display(room_id):
    """자습실 QR코드 표시 (인쇄용)"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    # QR 토큰이 없으면 생성
    if not room.qr_token:
        room.qr_token = secrets.token_hex(16)
        db.session.commit()

    # QR 코드가 가리킬 URL (입실 / 퇴실)
    qr_url         = url_for('student.qr_attend',   token=room.qr_token, _external=True)
    qr_checkout_url = url_for('student.qr_checkout', token=room.qr_token, _external=True)

    return render_template('teacher/qr_display.html',
                           room=room,
                           qr_checkout_url=qr_checkout_url,
                           qr_url=qr_url)


@teacher_bp.route('/settings/room/<int:room_id>/layout')
def seat_layout(room_id):
    """자습실 좌석 배치도 (드래그 앤 드롭 + 실시간 출결)"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    assigned = StudentRoom.query.filter_by(study_room_id=room_id).all()

    today = date.today()
    day_type = get_day_type(today)
    period_times = get_period_settings(day_type) if day_type != 'sunday' else {}

    return render_template('teacher/seat_layout.html',
                           room=room, assigned=assigned,
                           period_times=period_times, today=today)


@teacher_bp.route('/api/room/<int:room_id>/attendance_status')
def room_attendance_status(room_id):
    """자습실 학생 출결 현황 JSON API (날짜·교시 기준)"""
    date_str = request.args.get('date', date.today().isoformat())
    period_q  = request.args.get('period', type=int)   # None이면 전체 교시

    try:
        query_date = date.fromisoformat(date_str)
    except ValueError:
        return jsonify({'ok': False, 'error': '날짜 형식 오류'}), 400

    assigned = StudentRoom.query.filter_by(study_room_id=room_id).all()
    user_ids  = [sr.user_id for sr in assigned if sr.user_id]

    # 학생 N명에 대해 N+1 쿼리를 피하기 위해 한 번의 IN 쿼리로 모두 가져온다.
    atts_by_uid = {uid: {} for uid in user_ids}
    if user_ids:
        for a in Attendance.query.filter(
            Attendance.user_id.in_(user_ids),
            Attendance.date == query_date,
        ).all():
            atts_by_uid[a.user_id][a.period] = a

    apps_by_uid = {uid: set() for uid in user_ids}
    if user_ids:
        for sa in StudyApplication.query.filter(
            StudyApplication.user_id.in_(user_ids),
            StudyApplication.date == query_date,
        ).all():
            apps_by_uid[sa.user_id].add(sa.period)

    # 가장 좋은 상태를 대표값으로 — after_school(방과후출결인정)은 approved_leave와 동급
    PRIORITY = {'present': 0, 'late': 1, 'approved_leave': 2, 'after_school': 2,
                'early_leave': 3, 'applied': 4, 'absent': 5, 'none': 6}

    result = {}
    for uid in user_ids:
        atts = atts_by_uid.get(uid, {})
        apps = apps_by_uid.get(uid, set())

        def _status(p):
            if p in atts:
                return atts[p].status      # 'present' / 'late' / 'absent' / ...
            if p in apps:
                return 'applied'           # 신청했지만 아직 미처리
            return 'none'

        all_periods = sorted(set(atts.keys()) | apps)
        periods_data = {str(p): _status(p) for p in all_periods}

        if period_q is not None:
            overall = _status(period_q)
        else:
            stati = [_status(p) for p in all_periods] if all_periods else ['none']
            overall = min(stati, key=lambda s: PRIORITY.get(s, 9))

        result[str(uid)] = {'overall': overall, 'periods': periods_data}

    return jsonify({'ok': True, 'date': date_str, 'period': period_q, 'data': result})


@teacher_bp.route('/api/room/<int:room_id>/save_layout', methods=['POST'])
def save_layout(room_id):
    """좌석 배치도 위치 저장 (JSON API) — 모든 교사 공유"""
    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'ok': False, 'error': '데이터 없음'}), 400

    if not db.session.get(StudyRoom, room_id):
        return jsonify({'ok': False, 'error': '존재하지 않는 자습실'}), 404

    for item in data:
        try:
            uid = int(item['user_id'])
            x   = float(item['x'])
            y   = float(item['y'])
        except (KeyError, TypeError, ValueError):
            return jsonify({'ok': False, 'error': '잘못된 항목 데이터'}), 400
        if not (0.0 <= x <= 100.0) or not (0.0 <= y <= 100.0):
            return jsonify({'ok': False, 'error': '좌표는 0~100 범위여야 합니다'}), 400
        sr = StudentRoom.query.filter_by(user_id=uid, study_room_id=room_id).first()
        if sr:
            sr.pos_x = x
            sr.pos_y = y

    db.session.commit()
    return jsonify({'ok': True})


@teacher_bp.route('/qr/regenerate/<int:room_id>', methods=['POST'])
def qr_regenerate(room_id):
    """QR코드 토큰 재생성"""
    room = db.session.get(StudyRoom, room_id)
    if not room:
        flash('존재하지 않는 자습 공간입니다.', 'danger')
        return redirect(url_for('teacher.settings'))

    room.qr_token = secrets.token_hex(16)
    db.session.commit()

    flash(f'"{room.name}" QR코드가 재생성되었습니다.', 'success')
    return redirect(url_for('teacher.qr_display', room_id=room_id))


@teacher_bp.route('/mypage', methods=['GET', 'POST'])
@login_required
def mypage():
    """교사 마이페이지 - 비밀번호 변경"""
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not current_user.check_password(current_pw):
            flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
            return render_template('teacher/mypage.html')

        if new_pw != confirm_pw:
            flash('새 비밀번호가 일치하지 않습니다.', 'danger')
            return render_template('teacher/mypage.html')

        ok, err = validate_password(new_pw)
        if not ok:
            flash(err, 'danger')
            return render_template('teacher/mypage.html')

        current_user.set_password(new_pw)
        db.session.commit()
        flash('비밀번호가 변경되었습니다. 다시 로그인해 주세요.', 'success')
        return redirect(url_for('teacher.dashboard'))

    return render_template('teacher/mypage.html')
